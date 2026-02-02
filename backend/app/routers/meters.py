import csv
import io
import os
import time
import re
import math
from typing import Any, Optional, List, Dict

import httpx
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from pydantic import BaseModel, Field

from app.config import get_settings

router = APIRouter(prefix="/api/meters", tags=["meters"])

# Cache simple de tokens por concentrador (ip/base_url)
_TOKEN_CACHE: dict[str, dict[str, Any]] = {}
# Cache de mapeo de excel (se recarga si cambia el archivo)
_EXCEL_CACHE: dict[str, Any] = {"mtime": None, "meter_to_conc": {}, "conc_to_ip": {}}


class ReadReportIn(BaseModel):
    meter: str = Field(..., description="CIR del medidor o número (ej: 141825620)")
    report_name: str = Field(..., description="Ej: CIR7, S01, S02, S2B, S03, S04, S4E")
    priority: int = Field(2, ge=0, le=9)
    fini: Optional[str] = Field(None, description="ISO 8601, ej: 2026-01-24T00:01:00Z")
    fend: Optional[str] = Field(None, description="ISO 8601, ej: 2026-01-24T23:59:00Z")


class ReadOrderIn(BaseModel):
    meter: str = Field(..., description="Medidor (con o sin prefijo CIR)")
    order: int = Field(..., ge=0, le=1, description="0=corte (OPEN), 1=reconexión (CLOSE)")
    priority: int = Field(2, ge=0, le=5)
    fini: Optional[str] = Field(None, description="Fecha/hora de ejecución (ISO o STG-CD)")
    fend: Optional[str] = Field(None, description="Fecha/hora máxima de ejecución (ISO o STG-CD)")
    id_pet: int = Field(0, description="IdPet dentro del XML de orden")



def _normalize_cir(meter: str) -> tuple[str, int]:
    m = (meter or "").strip()
    if not m:
        raise HTTPException(status_code=400, detail="Debe indicar un medidor.")
    if m.upper().startswith("CIR"):
        cir = "CIR" + m[3:].strip()
        num = cir[3:]
    else:
        # si es numérico: lo convertimos a CIR + 10 dígitos (con cero a la izquierda)
        if not m.isdigit():
            raise HTTPException(status_code=400, detail="El medidor debe ser 'CIRxxxxxxxxxx' o un número.")
        num = str(int(m))  # elimina ceros a la izquierda si los hubiera
        num = num.zfill(10)
        cir = "CIR" + num

    # meter_id: entero sin el cero inicial (si lo hay)
    try:
        meter_id_int = int(cir[3:])  # int() ya quita ceros iniciales
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de medidor inválido.")
    return cir, meter_id_int

def _parse_meter_cell(v: Any) -> Optional[int]:
    """Normaliza un valor de Excel (int/float/str) a número de medidor (int).

    Evita el caso típico de Excel que devuelve 142414721.0 (float) y al limpiar no-dígitos
    quedaba '1424147210' (agregando un 0).
    """
    if v is None:
        return None

    # openpyxl puede devolver int/float, incluso si la celda era "numérica"
    try:
        import numpy as _np  # type: ignore
        if isinstance(v, (_np.integer,)):
            return int(v)
        if isinstance(v, (_np.floating,)):
            if _np.isnan(v):
                return None
            return int(round(float(v)))
    except Exception:
        pass

    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        if math.isnan(v):
            return None
        return int(round(v))

    s = str(v).strip()
    if not s:
        return None

    # Quita prefijo CIR
    s = s.replace("CIR", "").replace("cir", "").strip()

    # Si viene como '142414721.0' o notación científica, convertir a float->int
    try:
        if re.fullmatch(r"\d+\.\d+", s) or re.fullmatch(r"\d+(?:\.\d+)?[eE][+-]?\d+", s):
            return int(float(s))
    except Exception:
        pass

    # Si tiene punto, quedate con la parte entera ANTES del punto
    if "." in s:
        s = s.split(".", 1)[0]

    # Limpia todo lo que no sea dígito
    s = re.sub(r"\D", "", s)
    if not s:
        return None
    try:
        return int(s)
    except Exception:
        return None



def _load_excel_mapping():
    '''Carga mapeo desde concentradores.xlsx.

    Preferencia del cliente:
      - IP en fila 3
      - ID de concentrador en fila 9

    Para robustez, si la hoja no calza exactamente, se autodetectan filas:
      - ip_row: fila con mayor cantidad de celdas con patrón IP
      - conc_row: fila con mayor cantidad de IDs numéricos grandes (concentradores)
    '''
    import re as _re
    import openpyxl

    s = get_settings()
    xlsx_path = getattr(s, "concentradores_xlsx_path", None) or os.path.join(os.path.dirname(__file__), "..", "..", "data", "concentradores.xlsx")
    xlsx_path = os.path.abspath(xlsx_path)

    if not os.path.exists(xlsx_path):
        raise HTTPException(status_code=500, detail=f"No se encontró el archivo de concentradores: {xlsx_path}")

    mtime = os.path.getmtime(xlsx_path)
    if _EXCEL_CACHE["mtime"] == mtime:
        return

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Defaults solicitados
    default_ip_row = 3
    default_conc_row = 9
    first_data_col = 2  # normalmente la columna A contiene textos

    ip_regex = _re.compile(r"^\d{1,3}(?:\.\d{1,3}){3}$")

    conc_to_ip: dict[int, str] = {}
    meter_to_conc: dict[int, int] = {}

    def _detect_rows(ws):
        # ip_row: fila con más IPs
        best_ip = (0, default_ip_row)  # (score, row)
        for r in range(1, min(60, ws.max_row) + 1):
            score = 0
            for c in range(first_data_col, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                s = str(v).strip()
                if ip_regex.match(s):
                    score += 1
            if score > best_ip[0]:
                best_ip = (score, r)

        # conc_row: fila con más "concentradores" (números grandes)
        best_conc = (0, default_conc_row)  # (score, row)
        for r in range(1, min(80, ws.max_row) + 1):
            score = 0
            for c in range(first_data_col, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                try:
                    n = int(v)
                except Exception:
                    continue
                # Heurística: IDs de concentrador suelen ser grandes (>= 1e8)
                if n >= 100_000_000:
                    score += 1
            if score > best_conc[0]:
                best_conc = (score, r)

        # Elegir detectado si tiene suficientes evidencias, si no usar default
        ip_row = best_ip[1] if best_ip[0] >= 2 else default_ip_row
        conc_row = best_conc[1] if best_conc[0] >= 2 else default_conc_row

        if conc_row <= ip_row:
            # fallback a defaults si se detectó algo incoherente
            ip_row, conc_row = default_ip_row, default_conc_row

        return ip_row, conc_row

    for name in wb.sheetnames:
        ws = wb[name]
        ip_row, conc_row = _detect_rows(ws)

        # Columnas válidas: donde conc_row tenga un ID de concentrador.
        conc_by_col: dict[int, int] = {}
        for c in range(first_data_col, ws.max_column + 1):
            v = ws.cell(conc_row, c).value
            if v is None:
                continue
            try:
                conc_id = int(v)
            except Exception:
                continue
            conc_by_col[c] = conc_id

        if not conc_by_col:
            continue

        # Map concentrador->IP usando ip_row (misma columna).
        for c, conc_id in conc_by_col.items():
            ip_val = ws.cell(ip_row, c).value
            if ip_val is None:
                continue
            ip = str(ip_val).strip()
            conc_to_ip[conc_id] = ip

        # Map medidor->concentrador recorriendo filas desde conc_row+1
        for r in range(conc_row + 1, ws.max_row + 1):
            for c, conc_id in conc_by_col.items():
                v = ws.cell(r, c).value
                if v is None:
                    continue
                try:
                    meter_id = int(v)
                except Exception:
                    continue
                meter_to_conc[meter_id] = conc_id

    _EXCEL_CACHE["mtime"] = mtime
    _EXCEL_CACHE["meter_to_conc"] = meter_to_conc
    _EXCEL_CACHE["conc_to_ip"] = conc_to_ip


def _resolve_conc_and_ip_for_meter(meter_id_int: int) -> tuple[int, str]:
    _load_excel_mapping()
    meter_to_conc = _EXCEL_CACHE["meter_to_conc"]
    conc_to_ip = _EXCEL_CACHE["conc_to_ip"]

    conc_id = meter_to_conc.get(meter_id_int)
    if not conc_id:
        raise HTTPException(status_code=404, detail=f"No se encontró el medidor {meter_id_int} en concentradores.xlsx.")
    ip = conc_to_ip.get(conc_id)
    if not ip:
        raise HTTPException(status_code=404, detail=f"No se encontró IP para concentrador {conc_id} en concentradores.xlsx.")
    return conc_id, ip


def _resolve_ip_for_meter(meter_id_int: int) -> str:
    conc_id, ip = _resolve_conc_and_ip_for_meter(meter_id_int)
    return ip


async def _gede_login(base_url: str) -> str:
    # Cache por 10 minutos
    now = time.time()
    cached = _TOKEN_CACHE.get(base_url)
    if cached and cached.get("exp", 0) > now:
        return cached["token"]

    s = get_settings()
    username = getattr(s, "gede_username", "admin")
    password = getattr(s, "gede_password", "Adm1n")

    xml_body = f'<Login Username="{username}" Password="{password}"/>'
    url = base_url.rstrip("/") + "/login"

    async with httpx.AsyncClient(timeout=20) as client:
        r = await client.post(url, content=xml_body.encode("utf-8"), headers={"Content-Type": "application/xml"})

    if r.status_code not in (200, 201):
        raise HTTPException(status_code=502, detail=f"Login GEDE falló ({r.status_code}): {r.text[:300]}")

    # Parse XML de respuesta para extraer Token
    import xml.etree.ElementTree as ET
    try:
        root = ET.fromstring(r.text)
        token = root.attrib.get("Token") or root.attrib.get("token")
    except Exception:
        token = None

    if not token:
        raise HTTPException(status_code=502, detail="No se pudo leer el token del concentrador (respuesta de /login).")

    _TOKEN_CACHE[base_url] = {"token": token, "exp": now + 600}
    return token



async def _gede_logout(base_url: str, token: str) -> None:
    """Cierra la sesión en el concentrador para liberar recursos."""
    url = base_url.rstrip("/") + "/logout"
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            await client.post(url, headers={"Authorization": f"Bearer {token}"})
    except Exception:
        # best-effort: no romper el flujo si el logout falla
        pass


async def _gede_scale(base_url: str, token: str) -> None:
    """Escala privilegios del token actual (necesario para algunas órdenes como B03)."""
    url = base_url.rstrip("/") + "/scale"
    async with httpx.AsyncClient(timeout=20) as client:
        r = await client.post(url, headers={"Authorization": f"Bearer {token}"})
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Scale GEDE falló ({r.status_code}): {r.text[:300]}")


def _to_stg_ts(v: Optional[str]) -> Optional[str]:
    """Convierte un ISO-8601 (YYYY-MM-DDTHH:MM:SSZ) al formato STG-CD (YYYYMMDDHHMMSSmmmW)."""
    if not v:
        return None
    s = str(v).strip()
    if re.match(r"^\d{17}[A-Z]$", s):
        return s
    # Normalizar Z
    try:
        from datetime import datetime, timezone
        s2 = s.replace("Z", "+00:00")
        dt = datetime.fromisoformat(s2)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        dt = dt.astimezone(timezone.utc)
        return dt.strftime("%Y%m%d%H%M%S") + "000W"
    except Exception:
        return s



def _extract_eacti(data: Any) -> Optional[Any]:
    """Busca el campo Eacti (estado del relé) en distintos formatos."""
    if data is None:
        return None
    # Lista de filas (dicts)
    if isinstance(data, list) and data:
        row = data[0]
        if isinstance(row, dict):
            # keys posibles: 'Eacti' o variantes con prefijo
            if "Eacti" in row:
                return row.get("Eacti")
            for k, v in row.items():
                if k.lower().endswith("eacti"):
                    return v
    # Objeto dict
    if isinstance(data, dict):
        if "Eacti" in data:
            return data.get("Eacti")
        for k, v in data.items():
            if k.lower().endswith("eacti"):
                return v
    return None

def _try_parse_csv(text: str) -> Optional[list[dict[str, Any]]]:
    # intenta parsear CSV con separador ',' o ';'
    for delim in [",", ";", "	"]:
        try:
            f = io.StringIO(text)
            reader = csv.DictReader(f, delimiter=delim)
            rows = list(reader)
            if rows and reader.fieldnames and len(reader.fieldnames) >= 2:
                return rows
        except Exception:
            pass
    return None



def _strip_ns(tag: str) -> str:
    # '{ns}Tag' -> 'Tag'
    if not tag:
        return "Tag"
    if tag.startswith("{") and "}" in tag:
        return tag.split("}", 1)[1]
    return tag

def _xml_report_to_rows(xml_text: str) -> Optional[list[dict[str, Any]]]:
    """Convierte el XML de GEDE a filas/columnas.

    Estrategia:
      1) Buscar un 'record tag' (elemento hoja con atributos) repetido; si no hay repetidos,
         elegir el elemento hoja con MÁS atributos.
      2) Para cada record, generar una fila con:
         - atributos del root (Report)
         - atributos de ancestros (por ejemplo Cnc.Id, Cnt.Id)
         - atributos del record (sin prefijo si no colisiona; si colisiona, se prefija)
    """
    import xml.etree.ElementTree as ET
    try:
        root = ET.fromstring(xml_text)
    except Exception:
        return None

    # Recolectar candidatos hoja con atributos o texto
    leafs: list[ET.Element] = []
    all_elems: list[ET.Element] = []

    def walk(e: ET.Element):
        all_elems.append(e)
        children = list(e)
        if not children:
            leafs.append(e)
        for ch in children:
            walk(ch)

    walk(root)

    def leaf_score(e: ET.Element) -> int:
        return len(e.attrib or {}) + (1 if (e.text or "").strip() else 0)

    # Agrupar hojas por tag (sin namespace)
    groups: dict[str, list[ET.Element]] = {}
    for e in leafs:
        if leaf_score(e) == 0:
            continue
        t = _strip_ns(e.tag)
        groups.setdefault(t, []).append(e)

    # Elegir record group
    record_tag = None
    record_elems: list[ET.Element] = []

    # Preferir repetidos
    best = (0, 0)  # (count, avg_attr)
    for t, elems in groups.items():
        if len(elems) < 2:
            continue
        avg_attr = sum(len(x.attrib or {}) for x in elems) / max(1, len(elems))
        key = (len(elems), int(avg_attr))
        if key > best:
            best = key
            record_tag = t
            record_elems = elems

    if not record_elems:
        # Elegir el hoja con más atributos (o texto)
        best_elem = None
        best_score = -1
        for e in leafs:
            sc = leaf_score(e)
            if sc > best_score:
                best_score = sc
                best_elem = e
        if best_elem is None:
            return None
        record_tag = _strip_ns(best_elem.tag)
        record_elems = [best_elem]

    # Necesitamos ancestros: ElementTree no da parent, así que recorremos y armamos parent map
    parent: dict[ET.Element, ET.Element] = {}
    for e in all_elems:
        for ch in list(e):
            parent[ch] = e

    def ancestors(e: ET.Element):
        cur = e
        out = []
        while cur in parent:
            cur = parent[cur]
            out.append(cur)
        return out  # desde padre hacia root

    # Root attrs (Report)
    root_attrs = {}
    for k, v in (root.attrib or {}).items():
        root_attrs[_strip_ns(k)] = v

    rows: list[dict[str, Any]] = []
    for rec in record_elems:
        row: dict[str, Any] = dict(root_attrs)

        # Ancestros relevantes (hasta root)
        for a in ancestors(rec):
            tag = _strip_ns(a.tag)
            for k, v in (a.attrib or {}).items():
                col = f"{tag}.{_strip_ns(k)}"
                # no pisar si existe
                if col not in row:
                    row[col] = v

        # Record attrs
        used = set(row.keys())
        for k, v in (rec.attrib or {}).items():
            kk = _strip_ns(k)
            col = kk if kk not in used else f"{record_tag}.{kk}"
            row[col] = v

        # Texto del record si aplica
        t = (rec.text or "").strip()
        if t:
            col = "value" if "value" not in row else f"{record_tag}.value"
            row[col] = t

        # También incluir el tag del record
        if record_tag and "recordTag" not in row:
            row["recordTag"] = record_tag

        rows.append(row)

    return rows if rows else None

@router.post("/report")
async def read_report(payload: ReadReportIn):
    s = get_settings()
    cir, meter_id_int = _normalize_cir(payload.meter)

    conc_id, ip = _resolve_conc_and_ip_for_meter(meter_id_int)

    api_base = getattr(s, "gede_api_base", "/api/v1")
    base_url = f"http://{ip}{api_base}"

    token: Optional[str] = None
    try:
        token = await _gede_login(base_url)

        params = {
            "idMeters": cir,
            "priority": payload.priority,
        }
        if payload.fini:
            params["fini"] = payload.fini
        if payload.fend:
            params["fend"] = payload.fend


        url = base_url.rstrip("/") + f"/report/{payload.report_name}"

        try:
            async with httpx.AsyncClient(timeout=120) as client:
                r = await client.get(url, params=params, headers={"Authorization": f"Bearer {token}"})
        except httpx.ReadTimeout:
            raise HTTPException(
                status_code=504,
                detail=f"Timeout leyendo {payload.report_name} (IP {ip}). El concentrador no respondió a tiempo.",
            )

        # Si token expiró, reintenta una vez
        if r.status_code in (401, 403):
            _TOKEN_CACHE.pop(base_url, None)
            token = await _gede_login(base_url)
            try:
                async with httpx.AsyncClient(timeout=120) as client:
                    r = await client.get(url, params=params, headers={"Authorization": f"Bearer {token}"})
            except httpx.ReadTimeout:
                raise HTTPException(
                    status_code=504,
                    detail=f"Timeout leyendo {payload.report_name} (IP {ip}) en reintento. El concentrador no respondió a tiempo.",
                )

        if r.status_code != 200:
            raise HTTPException(status_code=502, detail=f"GEDE report falló ({r.status_code}): {r.text[:400]}")

        content_type = (r.headers.get("content-type") or "").lower()
        raw_text = r.text

        # Intentar JSON
        data: Any = None
        try:
            data = r.json()
        except Exception:
            data = None

        # Intentar CSV si no es JSON
        if data is None:
            parsed_csv = _try_parse_csv(raw_text)
            if parsed_csv is not None:
                data = parsed_csv

        # Intentar XML->filas si no es JSON ni CSV
        if data is None:
            parsed_xml = _xml_report_to_rows(raw_text)
            if parsed_xml is not None:
                data = parsed_xml

        return {
            "ip": ip,
            "conc_id": conc_id,
            "base_url": base_url,
            "report_name": payload.report_name,
            "meter": cir,
            "content_type": content_type,
            "data": data,
            "raw": raw_text,
            # "relay_eacti": relay_eacti,  # (NO) solo aplica a órdenes B03

        }
    finally:
        if token:
            await _gede_logout(base_url, token)
            _TOKEN_CACHE.pop(base_url, None)

@router.post("/order")
async def send_order(payload: ReadOrderIn):
    """Envía una orden B03 (corte/reconexión)."""
    s = get_settings()
    cir, meter_id_int = _normalize_cir(payload.meter)

    conc_id, ip = _resolve_conc_and_ip_for_meter(meter_id_int)

    api_base = getattr(s, "gede_api_base", "/api/v1")
    base_url = f"http://{ip}{api_base}"

    token: Optional[str] = None
    try:
        token = await _gede_login(base_url)

        # Escalado del token (requerido para B03 según Postman)
        await _gede_scale(base_url, token)

        from datetime import datetime, timezone, timedelta
        fini_ts = _to_stg_ts(payload.fini)
        fend_ts = _to_stg_ts(payload.fend)

        # UX B03: si el frontend envía una única fecha, usamos la misma para Fini y Ffin
        if fini_ts and not fend_ts:
            fend_ts = fini_ts

        if not fini_ts:
            fini_ts = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S") + "000W"
        if not fend_ts:
            fend_ts = (datetime.now(timezone.utc) + timedelta(days=1)).strftime("%Y%m%d%H%M%S") + "000W"

        xml_body = (
            f'<Order xmlns="http://stgdc/ws/B03" IdReq="B03" IdPet="{payload.id_pet}" Version="4.0">'
            f'<Cnc Id="CIR{conc_id}">'
            f'<Cnt Id="{cir}">'
            f'<B03 Fini="{fini_ts}" Ffin="{fend_ts}" Order="{payload.order}"/>'
            f'</Cnt></Cnc></Order>'
        )

        params = {"priority": payload.priority}
        url = base_url.rstrip("/") + "/order"

        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/xml"
        }

        async with httpx.AsyncClient(timeout=120) as client:
            r = await client.put(url, params=params, content=xml_body, headers=headers)
            if r.status_code == 405:
                r = await client.post(url, params=params, content=xml_body, headers=headers)

        # Si token expiró, reintenta una vez (incluye scale)
        if r.status_code in (401, 403):
            _TOKEN_CACHE.pop(base_url, None)
            token = await _gede_login(base_url)
            await _gede_scale(base_url, token)
            headers["Authorization"] = f"Bearer {token}"
            async with httpx.AsyncClient(timeout=120) as client:
                r = await client.put(url, params=params, content=xml_body, headers=headers)
                if r.status_code == 405:
                    r = await client.post(url, params=params, content=xml_body, headers=headers)

        if r.status_code != 200:
            raise HTTPException(status_code=502, detail=f"GEDE order B03 falló ({r.status_code}): {r.text[:400]}")

        raw_text = r.text
        content_type = (r.headers.get("content-type") or "").lower()

        data: Any = None
        try:
            data = r.json()
        except Exception:
            data = None

        if data is None:
            parsed_csv = _try_parse_csv(raw_text)
            if parsed_csv is not None:
                data = parsed_csv

        if data is None:
            parsed_xml = _xml_report_to_rows(raw_text)
            if parsed_xml is not None:
                data = parsed_xml


        # Luego de ejecutar B03, interrogamos S01 para leer el estado del relé (Eacti)
        relay_eacti = None
        try:
            import asyncio
            await asyncio.sleep(1.5)  # pequeña espera para que el estado se estabilice
            url_s01 = base_url.rstrip("/") + "/report/S01"
            params_s01 = {"idMeters": cir, "priority": payload.priority}
            async with httpx.AsyncClient(timeout=120) as client:
                r2 = await client.get(url_s01, params=params_s01, headers={"Authorization": f"Bearer {token}"})
            if r2.status_code == 200:
                raw2 = r2.text
                data2: Any = None
                try:
                    data2 = r2.json()
                except Exception:
                    data2 = None
                if data2 is None:
                    parsed_csv2 = _try_parse_csv(raw2)
                    if parsed_csv2 is not None:
                        data2 = parsed_csv2
                if data2 is None:
                    parsed_xml2 = _xml_report_to_rows(raw2)
                    if parsed_xml2 is not None:
                        data2 = parsed_xml2
                relay_eacti = _extract_eacti(data2)
        except Exception:
            relay_eacti = None

        return {
            "ip": ip,
            "conc_id": conc_id,
            "base_url": base_url,
            "report_name": "B03",
            "meter": cir,
            "order": payload.order,
            "content_type": content_type,
            "data": data,
            "raw": raw_text,
        }
    finally:
        if token:
            await _gede_logout(base_url, token)
            _TOKEN_CACHE.pop(base_url, None)


@router.post("/order_massive")
async def send_order_massive(
    order: int = Form(..., description="0=corte, 1=reconexion"),
    actdate: str = Form(..., description="Fecha ISO (ActDate)"),
    priority: int = Form(2),
    id_pet: int = Form(0),
    file: UploadFile = File(..., description="Excel con lista de medidores"),
):
    """Envía B03 masivo leyendo un Excel de medidores, y luego interroga S01 para obtener Eacti por cada uno.

    Devuelve una tabla con NIS/Nombre/Medidor/Estado para dar visibilidad de la tarea.
    """
    import openpyxl
    import asyncio

    # --- Cargar catálogo NIS/Nombre/Medidor (provisorio: Excel) ---
    s = get_settings()
    cat_path = getattr(s, "cadena_electrica_xlsx_path", None) or os.path.join(os.path.dirname(__file__), "..", "..", "data", "cadena_electrica_georreferenciacion.xlsx")
    cat_path = os.path.abspath(cat_path)

    cat_map: Dict[int, Dict[str, Any]] = {}
    if os.path.exists(cat_path):
        try:
            wb = openpyxl.load_workbook(cat_path, data_only=True)
            ws = wb.active
            # encabezados
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip() if cell.value is not None else "")
            low = [h.lower() for h in headers]

            def col_idx(*names):
                for n in names:
                    n = n.lower()
                    if n in low:
                        return low.index(n) + 1
                return None

            c_med = col_idx("medidor", "meter", "idmedidor", "cir")
            c_nis = col_idx("nis", "n.i.s", "nº nis", "numero nis")
            c_nom = col_idx("nombre", "name", "cliente")

            if c_med:
                for r in range(2, ws.max_row + 1):
                    v = ws.cell(row=r, column=c_med).value
                    mid = _parse_meter_cell(v)
                    if mid is None:
                        continue
                    cat_map[mid] = {
                        "nis": (str(ws.cell(row=r, column=c_nis).value).strip() if c_nis and ws.cell(row=r, column=c_nis).value is not None else None),
                        "nombre": (str(ws.cell(row=r, column=c_nom).value).strip() if c_nom and ws.cell(row=r, column=c_nom).value is not None else None),
                    }
        except Exception:
            cat_map = {}

    # --- Leer archivo subido con lista de medidores ---
    content = await file.read()
    try:
        wb2 = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el Excel subido. Verificá formato .xlsx/.xls.")

    ws2 = wb2.active

    # Detectar si la primera fila es encabezado (contiene texto con letras)
    first_row_vals = [c.value for c in ws2[1]]
    has_header = any(isinstance(v, str) and re.search(r"[A-Za-z]", v) for v in first_row_vals)

    col_m = 1
    start_row = 1

    if has_header:
        headers2 = []
        for cell in ws2[1]:
            headers2.append(str(cell.value).strip() if cell.value is not None else "")
        low2 = [h.lower() for h in headers2]

        for key in ("medidor", "meter", "idmedidor", "cir", "id"):
            if key in low2:
                col_m = low2.index(key) + 1
                break
        start_row = 2

    meters: List[int] = []
    for r in range(start_row, ws2.max_row + 1):
        v = ws2.cell(row=r, column=col_m).value
        mid = _parse_meter_cell(v)
        if mid is None:
            continue
        meters.append(mid)

    # dedup manteniendo orden
    seen = set()
    meters = [x for x in meters if not (x in seen or seen.add(x))]

    if not meters:
        raise HTTPException(status_code=400, detail="El archivo no contiene medidores válidos.")

    act_ts = _to_stg_ts(actdate)
    if not act_ts:
        raise HTTPException(status_code=400, detail="Fecha inválida (ActDate).")

    results: List[Dict[str, Any]] = []

    # --- secuencial para no saturar sesiones del concentrador ---
    for mid in meters:
        cir, mid_int = _normalize_cir(str(mid))
        conc_id = None
        ip = None
        base_url = None
        token: Optional[str] = None
        relay_eacti = None
        ok = False
        err = None

        try:
            conc_id, ip = _resolve_conc_and_ip_for_meter(mid_int)
            api_base = getattr(s, "gede_api_base", "/api/v1")
            base_url = f"http://{ip}{api_base}"

            token = await _gede_login(base_url)
            await _gede_scale(base_url, token)

            xml_body = (
                f'<Order xmlns="http://stgdc/ws/B03" IdReq="B03" IdPet="{id_pet}" Version="4.0">'
                f'<Cnc Id="CIR{conc_id}">'
                f'<Cnt Id="{cir}">'
                f'<B03 Fini="{act_ts}" Ffin="{act_ts}" Order="{order}"/>'
                f'</Cnt></Cnc></Order>'
            )

            url = base_url.rstrip("/") + "/order"
            params = {"priority": priority}
            headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/xml"}

            async with httpx.AsyncClient(timeout=120) as client:
                r = await client.put(url, params=params, content=xml_body, headers=headers)
                if r.status_code == 405:
                    r = await client.post(url, params=params, content=xml_body, headers=headers)

            if r.status_code != 200:
                raise Exception(f"B03 falló ({r.status_code}): {r.text[:200]}")

            # --- luego S01 para leer Eacti ---
            await asyncio.sleep(1.5)
            url_s01 = base_url.rstrip("/") + "/report/S01"
            params_s01 = {"idMeters": cir, "priority": priority}
            async with httpx.AsyncClient(timeout=120) as client:
                r2 = await client.get(url_s01, params=params_s01, headers={"Authorization": f"Bearer {token}"})

            if r2.status_code == 200:
                raw2 = r2.text
                data2: Any = None
                try:
                    data2 = r2.json()
                except Exception:
                    data2 = None
                if data2 is None:
                    parsed_csv2 = _try_parse_csv(raw2)
                    if parsed_csv2 is not None:
                        data2 = parsed_csv2
                if data2 is None:
                    parsed_xml2 = _xml_report_to_rows(raw2)
                    if parsed_xml2 is not None:
                        data2 = parsed_xml2
                relay_eacti = _extract_eacti(data2)

            ok = True

        except Exception as e:
            ok = False
            err = str(e)

        finally:
            if token and base_url:
                try:
                    await _gede_logout(base_url, token)
                except Exception:
                    pass
                _TOKEN_CACHE.pop(base_url, None)

        info = cat_map.get(mid_int, {})
        results.append({
            "nis": info.get("nis"),
            "nombre": info.get("nombre"),
            "medidor": mid_int,
            "accion": "corte" if order == 0 else "reconexion",
            "eacti": relay_eacti,
            "estado": ("Conectado" if str(relay_eacti) == "1" else "Desconectado" if str(relay_eacti) == "0" else None),
            "ok": ok,
            "error": err,
            "ip": ip,
            "concentrador": conc_id,
        })

    return {"count": len(results), "results": results}