from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel

import openpyxl


router = APIRouter(prefix="/api/tecnica", tags=["tecnica"])


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DATA_DIR = PROJECT_ROOT / "backend" / "data"
FACTURACION_XLSX = DATA_DIR / "Facturacion.xlsx"
USERS_JSON = DATA_DIR / "users_tecnica.json"


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _load_facturacion_rows() -> List[Dict[str, Any]]:
    """Carga Facturacion.xlsx a memoria.

    Se usa como fuente provisional hasta integrar Postgres.
    """
    if not FACTURACION_XLSX.exists():
        return []
    wb = openpyxl.load_workbook(str(FACTURACION_XLSX), data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    norm_headers = [_safe_str(h) for h in headers]

    out: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {}
        empty = True
        for c in range(1, ws.max_column + 1):
            key = norm_headers[c - 1]
            val = ws.cell(row=r, column=c).value
            if val is not None and _safe_str(val) != "":
                empty = False
            row[key] = val
        if not empty:
            out.append(row)
    return out


# Cache simple en memoria
_FACT_ROWS: Optional[List[Dict[str, Any]]] = None


def _get_fact_rows() -> List[Dict[str, Any]]:
    global _FACT_ROWS
    if _FACT_ROWS is None:
        _FACT_ROWS = _load_facturacion_rows()
    return _FACT_ROWS


def _norm_q(q: str) -> str:
    return " ".join(q.lower().strip().split())


def _match_row(row: Dict[str, Any], q: str) -> bool:
    # Campos comunes
    nis = _safe_str(row.get("NIS"))
    med = _safe_str(row.get("Medidor"))
    nombre = _safe_str(row.get("Nombre"))
    return q in _norm_q(nis) or q in _norm_q(med) or q in _norm_q(nombre)


@router.get("/lookup")
def lookup(query: str):
    q = _norm_q(query)
    if not q:
        raise HTTPException(status_code=400, detail="query vacío")

    rows = _get_fact_rows()
    matches = [r for r in rows if _match_row(r, q)]

    if not matches:
        return {"found": False, "matches": []}

    # Devuelve el primer match y además una lista acotada
    return {
        "found": True,
        "match": matches[0],
        "matches": matches[:25],
    }


class UserIn(BaseModel):
    nis: str
    nombre: str
    medidor: str
    direccion: str = ""
    tarifa: str = ""
    ruta: str = ""
    coordenada: str = ""


@router.post("/users")
def add_user(payload: UserIn):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    users: List[Dict[str, Any]] = []
    if USERS_JSON.exists():
        try:
            users = json.loads(USERS_JSON.read_text(encoding="utf-8"))
        except Exception:
            users = []

    users.append(payload.model_dump())
    USERS_JSON.write_text(json.dumps(users, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "count": len(users)}


@router.get("/users")
def list_users():
    if not USERS_JSON.exists():
        return {"users": []}
    try:
        users = json.loads(USERS_JSON.read_text(encoding="utf-8"))
    except Exception:
        users = []
    return {"users": users}
