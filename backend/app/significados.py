import os
from typing import Dict

from fastapi import HTTPException

from app.config import get_settings

_SIG_CACHE = {"mtime": None, "map": {}}  # type: ignore


def load_significados() -> Dict[str, str]:
    """Carga el diccionario de significados desde un XLSX.

    Estructura esperada (Hoja 1):
      Col A: Denominación (ej: Vf, L1v, Pimp, ...)
      Col B: Significado

    Cachea por mtime.
    """
    import openpyxl

    s = get_settings()
    xlsx_path = getattr(s, "significados_xlsx_path", None) or os.path.join(os.path.dirname(__file__), "..", "data", "Biblioteca Significados.xlsx")
    xlsx_path = os.path.abspath(xlsx_path)

    if not os.path.exists(xlsx_path):
        raise HTTPException(status_code=500, detail=f"No se encontró el archivo de significados: {xlsx_path}")

    mtime = os.path.getmtime(xlsx_path)
    if _SIG_CACHE["mtime"] == mtime and _SIG_CACHE["map"]:
        return _SIG_CACHE["map"]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    mapping: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 1).value
        meaning = ws.cell(r, 2).value
        if not code or not meaning:
            continue
        k = str(code).strip()
        v = str(meaning).strip()
        if not k or not v:
            continue
        mapping[k] = v

    _SIG_CACHE["mtime"] = mtime
    _SIG_CACHE["map"] = mapping
    return mapping
